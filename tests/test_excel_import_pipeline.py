"""Tests fuer die Excel-Import-Pipeline (Teile B-E des Daniel-Pilot-
Auftrags 2026-05-17).

Deckt:
- B.1  IF mit 2 Args (Excel-Kurzform, false-Branch=0)
- B.2  Label-Header anreichern (Spalten-Header + Zeilen-Label)
- B.3  Cross-Sheet-Refs (VK!B23) ueber alle Sheets aufloesen
- B.4  & String-Concat -> SKIP statt Crash
- C+D  Gemini-Classifier (Kategorisierung + Beschreibung, batched)
- E    Per-Entry-Wizard /ja /nein /skip

Externe Abhaengigkeiten (Gemini, DB) gemockt.
"""
from __future__ import annotations

import io
import uuid
from contextlib import asynccontextmanager
from types import SimpleNamespace
from unittest.mock import AsyncMock

import pytest

from core.integrations.excel_kalkulation import (
    SKIP_REASON_STRING_OP,
    _has_string_concat,
    _translate_excel_formula,
    extract_formulas_from_xlsx,
)


# =====================================================================
# Helper
# =====================================================================

def _make_xlsx(sheets: dict[str, list[list]]) -> bytes:
    """Baut Multi-Sheet-Xlsx aus {sheet_name: rows}."""
    import openpyxl
    wb = openpyxl.Workbook()
    # Default-Sheet entfernen
    default = wb.active
    wb.remove(default)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =====================================================================
# B.1: IF mit 2 Args
# =====================================================================

def test_if_with_2_args_uses_zero_as_else():
    """Excel-Kurzform IF(cond, val) -> false-branch = 0 (statt Crash)."""
    result = _translate_excel_formula(
        "=IF(A1=1, 100)",
        {"S1!A1": "stufenzahl"},
        current_sheet="S1",
    )
    # Erwartet: ((100) if (stufenzahl == 1) else (0))
    assert "stufenzahl" in result
    assert "100" in result
    assert "0" in result.split("else")[-1]


def test_if_with_3_args_still_works():
    """3-Arg-Form (vorher schon ok) muss weiter funktionieren."""
    result = _translate_excel_formula(
        "=IF(A1=1, 100, 200)",
        {"S1!A1": "stufenzahl"},
        current_sheet="S1",
    )
    assert "100" in result
    assert "200" in result


def test_if_with_4_args_still_errors():
    """4 Args bleiben ein Fehler (kein Excel-Konstrukt)."""
    with pytest.raises(ValueError, match="WENN/IF erwartet 2 oder 3"):
        _translate_excel_formula(
            "=IF(A1=1, 100, 200, 300)",
            {"S1!A1": "stufenzahl"},
            current_sheet="S1",
        )


def test_nested_if_outer_3args_inner_2args():
    """Daniels Pattern: aeusseres IF mit 3 Args, inneres mit 2 Args.

    Direkter Test auf _translate_excel_formula damit die DE/EN-Locale-
    Heuristik nicht reinpfuscht (xlsx-Round-Trip mit `1,2` wuerde als
    DE-Decimal interpretiert).
    """
    result = _translate_excel_formula(
        "=IF(A1=1, 100, IF(A1=2, 200))",
        {"S1!A1": "wahl"},
        current_sheet="S1",
    )
    assert "wahl" in result
    assert "100" in result
    assert "200" in result
    # innere IF ist 2-arg → else=0
    # Resultat: ((100) if (wahl==1) else (((200) if (wahl==2) else (0))))
    assert "0" in result.split("else")[-1]


# =====================================================================
# B.2: Label-Header (Spalten-Header + Zeilen-Label)
# =====================================================================

def test_block_table_combines_column_header_and_row_label():
    """Block-Layout: jede Summen-Zelle bekommt 'Header — Zeilen-Label'."""
    xlsx = _make_xlsx({
        "Treppe": [
            [None, "Standard", "Komfort", "Premium"],
            ["Material", 100, 150, 200],
            ["Summe", "=B2*1.2", "=C2*1.2", "=D2*1.2"],
        ],
    })
    result = extract_formulas_from_xlsx(xlsx)
    names = sorted(e.name for e in result.eintraege)
    assert "Standard — Summe" in names
    assert "Komfort — Summe" in names
    assert "Premium — Summe" in names
    # KEINE Suffix-Numerierung mehr noetig, weil Namen jetzt eindeutig
    assert "Summe #2" not in names
    assert "Summe #3" not in names


def test_label_falls_back_to_row_only_when_no_column_header():
    """Wenn nichts oben steht: nur Zeilen-Label."""
    xlsx = _make_xlsx({
        "S1": [
            ["entfernung_km", 50],
            ["preis", 2],
            ["Anfahrt", "=B1*B2"],
        ],
    })
    result = extract_formulas_from_xlsx(xlsx)
    assert len(result.eintraege) == 1
    # Spalten-Header von B3 wuerde B-Spalten-Header sein — in Zeile 1 ist
    # B1 = 50 (Number, kein Text-Header), in Zeile 2 = 2. Kein Header.
    # Row-Label = "Anfahrt"
    assert result.eintraege[0].name == "Anfahrt"


# =====================================================================
# B.3: Cross-Sheet-Refs
# =====================================================================

def test_cross_sheet_ref_resolves():
    """Maske!C5 = '=VK!B2*0.5' wird aufgeloest weil VK!B2 ein Label hat."""
    xlsx = _make_xlsx({
        "VK": [
            ["preis_standard", 100],
        ],
        "Maske": [
            ["Material-Aufschlag", "=VK!B1*1.2"],
        ],
    })
    result = extract_formulas_from_xlsx(xlsx)
    assert len(result.eintraege) == 1
    e = result.eintraege[0]
    assert "preis_standard" in e.variablen
    assert "preis_standard" in e.formel


def test_cross_sheet_ref_quoted_sheet_name():
    """Excel quoted Sheet-Names: 'Mein Sheet'!B1."""
    xlsx = _make_xlsx({
        "Mein Sheet": [
            ["stunden", 10],
        ],
        "Maske": [
            ["Lohn", "='Mein Sheet'!B1*75"],
        ],
    })
    result = extract_formulas_from_xlsx(xlsx)
    assert len(result.eintraege) == 1
    assert "stunden" in result.eintraege[0].variablen


def test_same_sheet_ref_still_works_without_prefix():
    """Same-Sheet-Refs (ohne Prefix) muessen weiter funktionieren."""
    xlsx = _make_xlsx({
        "S1": [
            ["km", 50],
            ["preis", 2],
            ["Anfahrt", "=B1*B2"],
        ],
    })
    result = extract_formulas_from_xlsx(xlsx)
    assert len(result.eintraege) == 1
    vars = result.eintraege[0].variablen
    assert "km" in vars and "preis" in vars


def test_cross_sheet_variable_names_no_collision():
    """Gleicher Label in 2 Sheets -> unique Variable-Namen (Slug-Dedup
    ueber alle Sheets)."""
    xlsx = _make_xlsx({
        "VK_A": [
            ["preis", 100],
        ],
        "VK_B": [
            ["preis", 200],
        ],
        "Maske": [
            ["Mix", "=VK_A!B1 + VK_B!B1"],
        ],
    })
    result = extract_formulas_from_xlsx(xlsx)
    assert len(result.eintraege) == 1
    # 2 unterschiedliche Variable-Namen erwartet
    assert len(result.eintraege[0].variablen) == 2


# =====================================================================
# B.4: &-Operator-Skip
# =====================================================================

def test_has_string_concat_detects_ampersand_outside_strings():
    assert _has_string_concat('=A1&":"') is True
    assert _has_string_concat('=A1&B1') is True


def test_has_string_concat_ignores_ampersand_inside_strings():
    """`&` innerhalb von Stringliteralen ist Beiwerk, nicht Operator."""
    assert _has_string_concat('="A & B"') is False


def test_string_concat_formula_is_skipped_with_dedicated_reason():
    """Excel-`&`-Formeln werden geskippt mit SKIP_REASON_STRING_OP."""
    xlsx = _make_xlsx({
        "Maske": [
            ["wahl", 1],
            ["Label", '=A1&":"'],
        ],
    })
    result = extract_formulas_from_xlsx(xlsx)
    assert result.eintraege == []
    assert result.verworfen_counts.get(SKIP_REASON_STRING_OP) == 1


# =====================================================================
# C+D: Gemini-Classifier
# =====================================================================

@pytest.mark.asyncio
async def test_classifier_returns_kategorie_and_beschreibung(monkeypatch):
    """Bei gueltiger Gemini-Antwort: pro Eintrag {kategorie, beschreibung}."""
    from core.ai import excel_classifier as cls
    fake_json = (
        '{"results": ['
        '{"index": 1, "kategorie": "pauschale", "beschreibung": "Standard-Treppe"},'
        '{"index": 2, "kategorie": "material", "beschreibung": "Material-Aufschlag"}'
        ']}'
    )
    monkeypatch.setattr(cls, "call_gemini", AsyncMock(return_value=fake_json))
    results = await cls.classify_excel_eintraege([
        {"name": "Standard — Summe", "formel": "a+b", "variablen": ["a", "b"], "sheet": "VK"},
        {"name": "Material", "formel": "c*1.2", "variablen": ["c"], "sheet": "VK"},
    ])
    assert len(results) == 2
    assert results[0]["kategorie"] == "pauschale"
    assert "Standard" in results[0]["beschreibung"]
    assert results[1]["kategorie"] == "material"


@pytest.mark.asyncio
async def test_classifier_strips_markdown_code_fences(monkeypatch):
    """Gemini wickelt manchmal JSON in ```json-Fences."""
    from core.ai import excel_classifier as cls
    fenced = (
        '```json\n'
        '{"results": [{"index": 1, "kategorie": "anfahrt", "beschreibung": "X"}]}\n'
        '```'
    )
    monkeypatch.setattr(cls, "call_gemini", AsyncMock(return_value=fenced))
    results = await cls.classify_excel_eintraege([
        {"name": "Anfahrt", "formel": "km*2", "variablen": ["km"], "sheet": "S"},
    ])
    assert results[0]["kategorie"] == "anfahrt"


@pytest.mark.asyncio
async def test_classifier_invalid_kategorie_falls_back_to_sonstiges(monkeypatch):
    """Wenn Gemini eine erfundene Kategorie zurueckgibt: 'sonstiges'."""
    from core.ai import excel_classifier as cls
    fake = '{"results": [{"index": 1, "kategorie": "erfunden", "beschreibung": "X"}]}'
    monkeypatch.setattr(cls, "call_gemini", AsyncMock(return_value=fake))
    results = await cls.classify_excel_eintraege([
        {"name": "X", "formel": "a", "variablen": ["a"], "sheet": "S"},
    ])
    assert results[0]["kategorie"] == "sonstiges"


@pytest.mark.asyncio
async def test_classifier_returns_fallback_on_gemini_error(monkeypatch):
    """Gemini-Crash: alle Eintraege als 'sonstiges' + leere Beschreibung."""
    from core.ai import excel_classifier as cls
    monkeypatch.setattr(
        cls, "call_gemini",
        AsyncMock(side_effect=RuntimeError("network down")),
    )
    results = await cls.classify_excel_eintraege([
        {"name": "A", "formel": "x", "variablen": ["x"], "sheet": "S"},
        {"name": "B", "formel": "y", "variablen": ["y"], "sheet": "S"},
    ])
    assert len(results) == 2
    assert all(r["kategorie"] == "sonstiges" for r in results)
    assert all(r["beschreibung"] == "" for r in results)


@pytest.mark.asyncio
async def test_classifier_returns_fallback_on_invalid_json(monkeypatch):
    """Gemini liefert kein JSON: fallback statt Crash."""
    from core.ai import excel_classifier as cls
    monkeypatch.setattr(
        cls, "call_gemini",
        AsyncMock(return_value="Sorry, ich kann das nicht."),
    )
    results = await cls.classify_excel_eintraege([
        {"name": "X", "formel": "a", "variablen": ["a"], "sheet": "S"},
    ])
    assert results[0]["kategorie"] == "sonstiges"


@pytest.mark.asyncio
async def test_classifier_empty_input_returns_empty_list():
    """Keine Eintraege -> kein Gemini-Call, leere Liste."""
    from core.ai import excel_classifier as cls
    results = await cls.classify_excel_eintraege([])
    assert results == []


@pytest.mark.asyncio
async def test_classifier_truncates_long_beschreibung(monkeypatch):
    """Beschreibung wird auf MAX_BESCHREIBUNG_LEN gekappt."""
    from core.ai import excel_classifier as cls
    long_text = "X" * 500
    fake = (
        '{"results": [{"index": 1, "kategorie": "sonstiges", '
        f'"beschreibung": "{long_text}"}}]}}'
    )
    monkeypatch.setattr(cls, "call_gemini", AsyncMock(return_value=fake))
    results = await cls.classify_excel_eintraege([
        {"name": "X", "formel": "a", "variablen": ["a"], "sheet": "S"},
    ])
    assert len(results[0]["beschreibung"]) <= cls.MAX_BESCHREIBUNG_LEN + 1


# =====================================================================
# E: Per-Entry-Wizard
# =====================================================================

@pytest.mark.asyncio
async def test_per_entry_wizard_ja_advances_and_marks_kept(monkeypatch):
    from plugins.telegram_notify import handler as tn
    saved_state = {}

    async def _fake_save(chat_id, key, data):
        saved_state["key"] = key
        saved_state["data"] = data

    monkeypatch.setattr(tn, "_save_state", _fake_save)
    state = {
        "eintraege": [
            {"name": "A", "formel": "x", "variablen": ["x"], "sheet": "S",
             "kategorie": "pauschale", "beschreibung": "A-Desc",
             "cell": "B1", "raw_excel": "=B1"},
            {"name": "B", "formel": "y", "variablen": ["y"], "sheet": "S",
             "kategorie": "material", "beschreibung": "B-Desc",
             "cell": "B2", "raw_excel": "=B2"},
        ],
        "filename": "test.xlsx",
        "index": 0,
        "decisions": [],
    }
    reply = await tn._handle_kalk_excel_per_entry_input(1, "ja", state)
    assert "Formel 2/2" in reply
    assert "B" in reply
    assert saved_state["data"]["decisions"] == [True]
    assert saved_state["data"]["index"] == 1


@pytest.mark.asyncio
async def test_per_entry_wizard_nein_advances_and_marks_skipped(monkeypatch):
    from plugins.telegram_notify import handler as tn
    saved_state = {}

    async def _fake_save(chat_id, key, data):
        saved_state["data"] = data

    monkeypatch.setattr(tn, "_save_state", _fake_save)
    state = {
        "eintraege": [
            {"name": "A", "formel": "x", "variablen": ["x"], "sheet": "S",
             "kategorie": "pauschale", "beschreibung": "",
             "cell": "B1", "raw_excel": "=B1"},
            {"name": "B", "formel": "y", "variablen": ["y"], "sheet": "S",
             "kategorie": "material", "beschreibung": "",
             "cell": "B2", "raw_excel": "=B2"},
        ],
        "filename": "test.xlsx",
        "index": 0,
        "decisions": [],
    }
    reply = await tn._handle_kalk_excel_per_entry_input(1, "nein", state)
    assert "Formel 2/2" in reply
    assert saved_state["data"]["decisions"] == [False]


@pytest.mark.asyncio
async def test_per_entry_wizard_last_entry_saves(monkeypatch):
    """Bei der letzten Formel: Speichern + Bilanz statt Naechste-Frage."""
    from plugins.telegram_notify import handler as tn
    added_rows: list = []

    class _StubSession:
        def add(self, obj):
            added_rows.append(obj)

        async def commit(self):
            pass

    @asynccontextmanager
    async def cm():
        yield _StubSession()

    monkeypatch.setattr(tn, "AsyncSessionLocal", cm)
    monkeypatch.setattr(tn, "_clear_state", AsyncMock())
    tenant = SimpleNamespace(id=uuid.uuid4())
    monkeypatch.setattr(tn, "_get_tenant_by_chat", AsyncMock(return_value=tenant))

    state = {
        "eintraege": [
            {"name": "A", "formel": "x", "variablen": ["x"], "sheet": "S",
             "kategorie": "pauschale", "beschreibung": "A-Beschreibung",
             "cell": "B1", "raw_excel": "=B1"},
        ],
        "filename": "test.xlsx",
        "index": 0,
        "decisions": [],
    }
    reply = await tn._handle_kalk_excel_per_entry_input(1, "ja", state)
    assert "1 Formel(n) gespeichert" in reply
    assert len(added_rows) == 1
    assert added_rows[0].kategorie == "pauschale"
    assert added_rows[0].beschreibung == "A-Beschreibung"


@pytest.mark.asyncio
async def test_per_entry_wizard_rejects_invalid_input(monkeypatch):
    """Unklare Antwort -> Hilfetext, kein State-Advance."""
    from plugins.telegram_notify import handler as tn
    state = {
        "eintraege": [
            {"name": "A", "formel": "x", "variablen": ["x"], "sheet": "S",
             "kategorie": "sonstiges", "beschreibung": "",
             "cell": "B1", "raw_excel": "=B1"},
        ],
        "filename": "test.xlsx",
        "index": 0,
        "decisions": [],
    }
    reply = await tn._handle_kalk_excel_per_entry_input(1, "vielleicht", state)
    assert "ja" in reply.lower()
    assert "skip" in reply.lower() or "nein" in reply.lower()


@pytest.mark.asyncio
async def test_per_entry_wizard_skipped_entries_not_in_db(monkeypatch):
    """Skipped Eintraege landen NICHT in der DB."""
    from plugins.telegram_notify import handler as tn
    added_rows: list = []

    class _StubSession:
        def add(self, obj):
            added_rows.append(obj)

        async def commit(self):
            pass

    @asynccontextmanager
    async def cm():
        yield _StubSession()

    saved_state = {}

    async def _fake_save(chat_id, key, data):
        saved_state["data"] = data

    monkeypatch.setattr(tn, "_save_state", _fake_save)
    monkeypatch.setattr(tn, "AsyncSessionLocal", cm)
    monkeypatch.setattr(tn, "_clear_state", AsyncMock())
    tenant = SimpleNamespace(id=uuid.uuid4())
    monkeypatch.setattr(tn, "_get_tenant_by_chat", AsyncMock(return_value=tenant))

    state = {
        "eintraege": [
            {"name": "A", "formel": "x", "variablen": ["x"], "sheet": "S",
             "kategorie": "sonstiges", "beschreibung": "",
             "cell": "B1", "raw_excel": "=B1"},
            {"name": "B", "formel": "y", "variablen": ["y"], "sheet": "S",
             "kategorie": "sonstiges", "beschreibung": "",
             "cell": "B2", "raw_excel": "=B2"},
        ],
        "filename": "test.xlsx",
        "index": 0,
        "decisions": [],
    }
    # Erste Formel skippen
    await tn._handle_kalk_excel_per_entry_input(1, "skip", state)
    # State wurde geupdatet — naechste Antwort baut darauf
    new_state = saved_state["data"]
    # Zweite Formel behalten
    reply = await tn._handle_kalk_excel_per_entry_input(1, "ja", new_state)
    assert "1 Formel(n) gespeichert" in reply
    assert "1 uebersprungen" in reply
    assert len(added_rows) == 1
    assert added_rows[0].name == "B"
